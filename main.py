#Reading the xlsx document
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from math import floor
arq = openpyxl.load_workbook('planilhas.xlsx') 
matricula_aluno = int(67106); #matricula do aluno

#Estrutura do arquivo xlsx
#Geral: Identificação | Professor | Período | Peso | Matéria
#'Nome_d_matéria': Matrícula | Nome | Avaliação 1 | Peso A1 | Avaliação 2 | Peso A2...

#Estrutura do Data_base: { //Dicionário
# 'Nome' : DANIEL JACOB TONN OK 
# 'Matricula' : 2019000000   OK
# 'Estatística': { //Dicionário
#                   'Período': 2022.2
#                   'Peso': 2       
#                   'Nota': (7.0) //Tupla
#                   'Avaliação1': (7.0, 2) //Tupla
#                   'Avaliação2': (8.0, 2) //Tupla
#                          }
# 'Análise Real': { //Dicionário
#                   'Período': 2022.2
#                   'Peso': 2       
#                   'Nota': (7.0) //Tupla
#                   'Avaliação1': (7.0, 2) //Tupla
#                   'Avaliação2': (8.0, 2) //Tupla
#                          }
# }


#Estrutura do CR: { //Dicionário
# 'Acumulado': 7.0
# 'Período2022.2': 7.0
# 'Período2022.1': 7.0
# 'Período2021.2': 7.0
# }


def build(matricula, arq ): #function to find the subjects of the student
    Data_base = {}; #dicionario com as informações
    Data_base['Matrícula'] = matricula_aluno; 
    geral = pd.read_excel('planilhas.xlsx', sheet_name=0)#sheet with the general information
    for i in range(1, len(arq.sheetnames)): #for each sheet
        nome_materia = '';#nome das matérias 
        nota_materia = 0;# nota das matérias 
        peso_materia = 0; # peso das matérias
        nota_por_avaliacao = []; #lista com as notas das avaliações
        dados_materia = {}; #dicionario com as informações da matéria
        planilha = pd.read_excel('planilhas.xlsx', sheet_name=i) #read the sheet
        for j in range(0, len(planilha)): #for each line in sheet
            if planilha.iloc[j, 0] == matricula: #if the student is in the sheet
                Data_base['Nome'] = planilha.iloc[j, 1] #add the name to the dictionary
                nota = 0;
                for k in range(0, len(geral)): #for each line in the general sheet
                    if (geral.iloc[k, 0] == arq.sheetnames[i]): #find the name of the subject
                        nome_materia = geral.iloc[k]['Matéria']; #add the name of the subject to the list
                        dados_materia['Período'] = geral.iloc[k]['Período']; #add the period to the dictionary
                        dados_materia['Peso'] = geral.iloc[k]['Peso'] #add the weight of the subject to the list
                        break; #break for k
                soma_pesos = 0; #sum of the weights
                for k in range(2, len(planilha.columns)): #for each column in the sheet
                    if(k%2 == 0):
                        soma_pesos += planilha.iloc[j, k+1]; #add the weight to the sum
                        nota_por_avaliacao.append((planilha.iloc[j, k], planilha.iloc[j, k+1])) #add the grade and the weight of the grade to the list
                        nota += planilha.iloc[j, k]*planilha.iloc[j, k+1] #add the grade to the total grade
                nota = nota/soma_pesos; #calculate the final grade
                dados_materia['Nota'] = nota; #add the final grade to the dictionary
                for k in range(0, len(nota_por_avaliacao)):
                    dados_materia['Avaliação'+str(k+1)] = nota_por_avaliacao[k];
                break; #break for j
        Data_base[nome_materia] = dados_materia; #add the dictionary to the dictionary
    return Data_base; #return the dictionary

def calculateCR(Data_base): #function to calculate the CR
    soma_pesos = 0; #sum of the weights
    soma_notas = 0; #sum of the grades
    CR = {}
    periodos_cursados = [];
    print(Data_base);
    for i in Data_base: #for each subject
        if(i != 'Nome' and i != 'Matrícula'): #if the subject is not the name or the registration number
            soma_pesos += Data_base[i]['Peso']; #add the weight to the sum
            soma_notas += Data_base[i]['Peso']*Data_base[i]['Nota']; #add the grade to the sum
            if Data_base[i]['Período'] not in periodos_cursados:
                periodos_cursados.append(Data_base[i]['Período']);
    CR['Acumulado'] = soma_notas/soma_pesos; #calculate the CR
    for i in periodos_cursados: #for each subject
        soma_pesos = 0; #sum of the weights
        soma_notas = 0; #sum of the grades
        for j in Data_base: #for each subject
            if(j!= 'Nome' and j != 'Matrícula' and Data_base[j]['Período'] == i):
                soma_pesos += Data_base[j]['Peso'];
                soma_notas += Data_base[j]['Peso']*Data_base[j]['Nota'];
        CR['Período'+str(i)] = soma_notas/soma_pesos;
    return CR; #return the CR

def plot(Date_base):
    CR = calculateCR(Date_base);
    #encontra quais são os períodos
    periodos = [];
    for i in Date_base:
        if(i != 'Nome' and i != 'Matrícula'):
            if Date_base[i]['Período'] not in periodos:
                periodos.append(Date_base[i]['Período']);
    #separa as matérias por período
    materias = [];
    for i in periodos: #para cada período
        materias.append([]);#adiciono uma lista vazia no período
    for i in Date_base: #para cada matéria
        if(i != 'Nome' and i != 'Matrícula'):
            for j in range(0, len(periodos)): #para cada período
                if Date_base[i]['Período'] == periodos[j]:#se periodo da materia == periodo
                    materias[j].append(i); #adiciono a matéria na lista desse período
    #define space between the bars
    peso_de_cada_barra = [];
    for i in materias: #para cada período
            peso_de_cada_barra.append([]); #adiciono uma lista vazia
    for k in Data_base: #para cada matéria geral
            if(k != 'Nome' and k != 'Matrícula'):
                    for l in range(0, len(periodos)): #se o periodo é igual ao periodo[l];
                     if( Data_base[k]['Período'] == periodos[l]):
                        peso_de_cada_barra[l].append(Date_base[k]['Peso']);
    space = [];
    for i in range(0, len(peso_de_cada_barra)): #para cada periodo, adiciono uma lista
        space.append([]);
    for k in range(0, len(space)): #para cada periodo
        soma = 0; #espaço inicial é zero inicialmente
        for i in range(0, len(peso_de_cada_barra[k])): #para cada matéria
                if(i == 0):#se for a primeira matéria
                    space[k].append(peso_de_cada_barra[k][0]/2); #adiciono o peso da primeira matéria dividido por 2
                    soma = peso_de_cada_barra[k][0] + 0.5;
                else:
                    soma = soma + peso_de_cada_barra[k][i]/2;
                    space[k].append(soma);
                    soma = soma + peso_de_cada_barra[k][i]/2 + 0.5;
    #encontro nota de cada materia
    notas = [];
    pesos_individuais = [];
    notas_individuais = [];
    cores = [];
    for i in range(0, len(materias)):
        notas.append([]);
        pesos_individuais.append([]);
        notas_individuais.append([]);
        cores.append([]);
        for j in range(0,len(materias[i])):
            pesos_individuais[i].append([]); #adiciono uma lista vazia
            notas_individuais[i].append([]);
            cores[i].append([]);
            notas[i].append(Data_base[materias[i][j]]['Nota']);

         #include the fist element of each avaliação in each materia inthe list pesos_individuais
    for i in Data_base: #pra cada matéria 
        if(i != 'Nome' and i != 'Matrícula'):
            for j in Data_base[i]:
                if (j != 'Peso' and j != 'Período' and j != 'Nota'):
                    #até aqui selecionei quem são as avaliações: Data_base[i][j]
                    for k in range(0, len(pesos_individuais)):
                        for l in range(0, len(pesos_individuais[k])):
                            #aqui cheguei na lista vazia de cada materia: pesos_individuais[k][l]
                            if(i == materias[k][l]):
                                pesos_individuais[k][l].append(Data_base[i][j][1]);
                                notas_individuais[k][l].append(Data_base[i][j][0]);
                                if Data_base[i][j][0] >= 7:
                                    cores[k][l].append('green');
                                elif Data_base[i][j][0] >= 6:
                                    cores[k][l].append('yellow');
                                else:
                                    cores[k][l].append('red');
                                
    #PLOT PLOT PLOT PLOT PLOT

    data_set = [];
    for i in range(0, len(periodos)):
        aux = [];
        aux = [space[i], notas[i], peso_de_cada_barra[i]];
        data_set.append(aux);
    #plot the graphs
    #plot two xy axis in the same plot
    linhas = 0
    if len(periodos) % 2 == 0:
        linhas = len(periodos)/2;
    else:
        linhas = (len(periodos)+1)/2;
    linhas = int(linhas);
    fig, axs = plt.subplots(linhas,2);
    fig.suptitle(Data_base['Nome']+'-'+str(Data_base['Matrícula']));
    for i in range(0, len(periodos)):
        a = int(floor(i/2));
        b = int(i%2);
        axs[a,b].bar(data_set[i][0], data_set[i][1], data_set[i][2]);
        axs[a,b].set_title('Período '+str(periodos[i]));
        axs[a,b].set_xticks(data_set[i][0], materias[i], rotation=45);
        #limit the y axis
        #show CR
        axs[a,b].hlines(CR['Acumulado'], 0, 10, color='violet', linestyle='--');
        axs[a,b].hlines(CR['Período'+str(periodos[i])], 0, 10, color='lightgreen', linestyle='--');
        #show CR as label
        #show CR in the left side and CR acumluda in the right side
        #show legend with CR
        axs[a,b].legend(['CR Acumulado', 'CR'+str(periodos[i])]);
        #axs[a,b].text(0.5, 0.5, 'CR: '+str(round(CR['Acumulado'],2)), transform=axs[a,b].transAxes, color = 'blue');
        #axs[a,b].text(0.5, 0.5, 'CR: '+str(round(CR['Período'+str(periodos[i])], 2)), transform=axs[a,b].transAxes, color = 'green');
        axs[a,b].set_ylim(0,10);
        axs[a,b].set_xlim(0,10);
        for j in range(0, len(space[i])):
                axs[a,b].text(space[i][j], notas[i][j], str(round(notas[i][j],2)), color='black', ha='center', va='bottom');
        for j in range(0, len(pesos_individuais[i])):#pra cada materia
            baixo = 0;
            peso_por_materia  = 0;
            for k in range(0, len(pesos_individuais[i][j])): #defino o peso total da disciplina
                peso_por_materia += pesos_individuais[i][j][k]; 
            for k in range(0, len(pesos_individuais[i][j])): #pra cada nota individual
                axs[a,b].bar(space[i][j], pesos_individuais[i][j][k]*notas[i][j]/peso_por_materia, peso_de_cada_barra[i][j], bottom=baixo, color = cores[i][j][k]); 
                if(k!=0):
                    axs[a,b].plot([space[i][j]-peso_de_cada_barra[i][j]/2, space[i][j]+peso_de_cada_barra[i][j]/2], [baixo, baixo], color='white', linewidth=1.5)
                    #axs[a,b].text(space[i][j], baixo, str(round(baixo, 2)), fontsize=15, color='black') 
                baixo = baixo + pesos_individuais[i][j][k]*notas[i][j]/peso_por_materia;
                    #    plt.plot([space[i]-weights_materia[i]*0.5, space[i]+weights_materia[i]*0.5], [baixo, baixo], color='white', linewidth=1.5)
                # #plot the grades for each avaliation
          #plot the grades for each avaliatio # 
                #pesos_individuais[i][j][k] = pesos_individuais[i][j][k]/soma;
                # #plot the grades for each avaliation
                # plt.text(space[i], baixo, str(round(individual[i][j], 2)), fontsize=15, color='white') 
                # baixo = baixo +   individual[i][j+1]*nota_materia[i]/pesospormateria;
                # #plot line above rectangles
                # plt.plot([space[i]-weights_materia[i]*0.5, space[i]+weights_materia[i]*0.5], [baixo, baixo], color='white', linewidth=1.5)
                # #plot the grades for each avaliation
    













       
        # individual = [];
        # for i in range(0, len(materias)):
        #     individual.append([]);
        #     for j in range(0, len(materias[i])):
        #         individual[i].append([]);
        # for i in Data_base:
        #     if i != 'Nome' and i != 'Matrícula':
        #         for j in (Data_base[i]):
        #             if j!= 'Periodo' and j!= 'Peso' and j!= 'Nota':
        #                 for k in range(0, len(periodos))
        #                     if Data_base[i]['Período'] == periodos[j]:
        #                         individual[j].append(Data_base[i]['Nota']);
        #                         individual[j].append(Data_base[i]['Peso']);
        # for i in range(0,len(individual)):
        # baixo = 0;
        # pesospormateria = 0;
        # for j in range(0, len(individual[i])):
        #     if(j%2 == 0):
        #         pesospormateria += individual[i][j+1];
        # cores='';
        # for j in range(0, len(individual[i])):
        #     if(j%2==0):
        #         if(individual[i][j] < 6):
        #             cores = 'r';
        #         elif(individual[i][j] < 7):
        #             cores = 'y';
        #         else:
        #             cores = 'g';
        #         plt.bar(space[i], individual[i][j+1]*nota_materia[i]/pesospormateria, weights_materia[i], bottom=baixo, color=cores); 
        #         #plot the grades for each avaliation
        #         plt.text(space[i], baixo, str(round(individual[i][j], 2)), fontsize=15, color='white') 
        #         baixo = baixo +   individual[i][j+1]*nota_materia[i]/pesospormateria;
        #         #plot line above rectangles
        #         plt.plot([space[i]-weights_materia[i]*0.5, space[i]+weights_materia[i]*0.5], [baixo, baixo], color='white', linewidth=1.5)
        #         #plot the grades for each avaliation























    fig.tight_layout()
    plt.show();



    
 






# for i in range(1, len(arq.sheetnames)): #for each planilha in the xlsx document
#         planilha = pd.read_excel('planilhas.xlsx', sheet_name=i) #read the planilha
#         for j in range(0, len(planilha)-1): #forline in planilha
#             if(planilha.iloc[j, 0] == matricula_aluno): #if matricula_aluno is in the planilha
#                 nome_materia.append(arq.sheetnames[i]); #add the name of the materia in the list
#             nota_final = 0;
#             soma_pesos = 0;
#             weightPgrades = [];
#             for k in range(2, (len(planilha.columns))): #for each avaliation in planilha
#                     weightPgrades.append(planilha.iloc[j, k]); 
#                     if (k % 2 == 0): #if the column is a weight
#                         soma_pesos = soma_pesos + planilha.iloc[j, k+1]; #sum the weights_materia
#                         nota_final = nota_final + planilha.iloc[j,k]*planilha.iloc[j,k+1]; #add the nota * peso
#             individual.append(weightPgrades); #add list of avaliations in the list
#             nota_materia.append(nota_final/soma_pesos); #add the nota_final of the materia in the list
#             break;
################################
 #We have the name of the materia and the nota of the materia,
 # so height of the bar and how many bar there is
################################


#chamar função findmaterias
Data_base = build(matricula_aluno, arq);
plot(Data_base);


# #PESOS DAS MATERIAS
# #defining the weights_materia of the bars
# weights_materia = [];
# planilha = pd.read_excel('planilhas.xlsx', sheet_name='Geral');
# for i in range(0, len(nome_materia)):
#     for j in range(0, len(planilha)):
#         if(planilha.iloc[j, 0] == nome_materia[i]):
#             weights_materia.append(planilha.iloc[j, 2]);
  


# #CR DO PERÍODO
# #media ponderada de weighs e nota_materia
# media_ponderada = 0;
# for i in range(0, len(weights_materia)):
#     media_ponderada += weights_materia[i]*nota_materia[i];
# media_ponderada = media_ponderada/sum(weights_materia);




# #define space between bars
# space = [];
# acumulado = 0;
# space_around = 0.1;
# for i in range(0, len(weights_materia)):
#     if(i==0):
#         space.append(0.5*weights_materia[i] + space_around);
#         acumulado = acumulado + weights_materia[i] + space_around;
#     else:
#         space.append(acumulado + weights_materia[i]*0.5 + space_around);
#         acumulado = acumulado + weights_materia[i]+ space_around;

# #PERÍODO
# #extracting numbers from a nome_materia
# periodo = '';
# p = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '.'];
# for i in range(0, len(nome_materia[0])):
#     if(nome_materia[0][i] in p):
#         periodo = periodo + nome_materia[0][i];


# #NOME DA MATÉRIA
# #deleting numbers and symbols in nome_materia
# for i in range(0, len(nome_materia)):
#     nome_materia[i] = nome_materia[i].replace("1", "");
#     nome_materia[i] = nome_materia[i].replace("2", "");
#     nome_materia[i] = nome_materia[i].replace("3", "");
#     nome_materia[i] = nome_materia[i].replace("4", "");
#     nome_materia[i] = nome_materia[i].replace("5", "");
#     nome_materia[i] = nome_materia[i].replace("6", "");
#     nome_materia[i] = nome_materia[i].replace("7", "");
#     nome_materia[i] = nome_materia[i].replace("8", "");
#     nome_materia[i] = nome_materia[i].replace("9", "");
#     nome_materia[i] = nome_materia[i].replace("0", "");
#     nome_materia[i] = nome_materia[i].replace("-", "");
#     nome_materia[i] = nome_materia[i].replace(".", "");

# # plot a histogram
# plt.bar(space, nota_materia, weights_materia);
# plt.yticks([]);
# plt.xticks([]);
# plt.xticks(space, nome_materia);
# plt.title(periodo)
# #add a vertical line in the mean of the notas
# #to slope xticks 
# plt.xticks(rotation=45);



# #LINHA NOTA MÍNIMA 6 E NOTA MÁXIMA 10
# plt.axhline(6, color='b', linestyle='dashed', linewidth=1)
# plt.axhline(10, color='black', linestyle='solid', linewidth=1)
# plt.axhline(media_ponderada, color='g', linestyle='dashed', linewidth=1)
# #plot media_ponderada in the y axis
# plt.text(0, media_ponderada, 'CR: ' + str(round(media_ponderada, 2)), fontsize=6, color='g')
# plt.text(0, 6, str(6), fontsize=6, color='b')


# for i in range(0,len(individual)):
#     baixo = 0;
#     pesospormateria = 0;
#     for j in range(0, len(individual[i])):
#         if(j%2 == 0):
#             pesospormateria += individual[i][j+1];
#     cores='';
#     for j in range(0, len(individual[i])):
#         if(j%2==0):
#             if(individual[i][j] < 6):
#                 cores = 'r';
#             elif(individual[i][j] < 7):
#                 cores = 'y';
#             else:
#                 cores = 'g';
#             plt.bar(space[i], individual[i][j+1]*nota_materia[i]/pesospormateria, weights_materia[i], bottom=baixo, color=cores); 
#             #plot the grades for each avaliation
#             plt.text(space[i], baixo, str(round(individual[i][j], 2)), fontsize=15, color='white') 
#             baixo = baixo +   individual[i][j+1]*nota_materia[i]/pesospormateria;
#             #plot line above rectangles
#             plt.plot([space[i]-weights_materia[i]*0.5, space[i]+weights_materia[i]*0.5], [baixo, baixo], color='white', linewidth=1.5)
#             #plot the grades for each avaliation





# plt.show();



